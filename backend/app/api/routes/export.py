"""Excel export API route."""

import io

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel

router = APIRouter()


class ExportRequest(BaseModel):
    """Full app state for Excel export."""
    profile: dict
    income: dict | None = None
    allocation: dict | None = None
    withdrawal: dict | None = None
    property_data: dict | None = None


@router.post("/export-excel")
async def export_excel(request: ExportRequest):
    """Generate .xlsx from full app state."""
    wb = Workbook()

    # Profile sheet
    ws = wb.active
    ws.title = "Profile"
    if request.profile:
        ws.append(["Field", "Value"])
        for key, value in request.profile.items():
            if key != "validationErrors":
                ws.append([key, str(value)])

    # Income sheet
    if request.income:
        ws_income = wb.create_sheet("Income")
        ws_income.append(["Field", "Value"])
        for key, value in request.income.items():
            if key != "validationErrors":
                ws_income.append([key, str(value)])

    # Allocation sheet
    if request.allocation:
        ws_alloc = wb.create_sheet("Allocation")
        ws_alloc.append(["Field", "Value"])
        for key, value in request.allocation.items():
            if key != "validationErrors":
                ws_alloc.append([key, str(value)])

    # Withdrawal sheet
    if request.withdrawal:
        ws_wd = wb.create_sheet("Withdrawal")
        ws_wd.append(["Field", "Value"])
        for key, value in request.withdrawal.items():
            if key != "validationErrors":
                ws_wd.append([key, str(value)])

    # Property sheet
    if request.property_data:
        ws_prop = wb.create_sheet("Property")
        ws_prop.append(["Field", "Value"])
        for key, value in request.property_data.items():
            if key != "validationErrors":
                ws_prop.append([key, str(value)])

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=fireplanner-export.xlsx"},
    )
